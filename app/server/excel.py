"""Excel surfaces (D2/D3) — the actuary's home turf, but carrying its provenance.

D3 · committee pack: a one-click .xlsx of the signed ultimates + the selection each
rests on + the audit trail, so the workbook that leaves the platform still carries the
selection_id, approver and rationale behind every number. "Your Excel now has lineage."

D2 · write-back template: a .xlsx an analyst fills in (line, factors, rationale) and the
app reads back to land a PENDING_APPROVAL selection — the SAME governed row the app /
notebook / SQL / MCP write. Excel becomes another door to the one governed record, not a
parallel unversioned copy.

openpyxl only; everything is built in-memory and streamed, nothing written to disk.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- house style, so the pack looks like a committee document, not a CSV dump ---
_NAVY = "1E293B"
_BLUE = "2563EB"
_LGREY = "F1F5F9"
_GREEN = "16A34A"
HDR = Font(bold=True, color="FFFFFF", size=11)
TITLE = Font(bold=True, size=15, color=_NAVY)
SUB = Font(italic=True, size=9, color="64748B")
BOLD = Font(bold=True)
MONEY = "#,##0"
PCT = "0.0%"
_thin = Side(style="thin", color="CBD5E1")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr_fill():
    return PatternFill("solid", fgColor=_NAVY)


def _band():
    return PatternFill("solid", fgColor=_LGREY)


def _autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title(ws, title, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, title).font = TITLE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(2, 1, sub).font = SUB


def _header_row(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = HDR; cell.fill = _hdr_fill(); cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def committee_pack(signoff, selections, audit, valuation_date, entity):
    """Build the committee .xlsx. Returns bytes.
    signoff:   rows with line_of_business_code, signed_best_estimate, reserving_method_code,
               selection_id, data_version, status_code, signed_by
    selections: the approved selected_development_pattern rows (id, factors, rationale, approver)
    audit:     recent audit events (created_at, event_type, entity_id, detail, actor)
    """
    wb = Workbook()

    # sheet 1 — signed reserves (the number + who signed + on what basis)
    ws = wb.active; ws.title = "Signed reserves"
    _title(ws, f"{entity} — Reserving Committee Pack",
           f"Valuation {valuation_date} · generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC · "
           f"every number carries the selection, approver and data version behind it", 7)
    hdrs = ["Line of business", "Signed best estimate", "Method", "Selection ID",
            "Data version", "Status", "Signed by"]
    _header_row(ws, 4, hdrs)
    r = 5
    for s in signoff:
        ws.cell(r, 1, (s.get("line_of_business_code") or "").replace("_", " "))
        c2 = ws.cell(r, 2, float(s.get("signed_best_estimate") or 0)); c2.number_format = MONEY
        ws.cell(r, 3, s.get("reserving_method_code"))
        ws.cell(r, 4, s.get("selection_id") or "—")
        ws.cell(r, 5, s.get("data_version") or "—")
        ws.cell(r, 6, (s.get("status_code") or "").replace("_", " ").title())
        ws.cell(r, 7, s.get("signed_by") or "—")
        for c in range(1, 8):
            ws.cell(r, c).border = BORDER
            if r % 2 == 0:
                ws.cell(r, c).fill = _band()
        r += 1
    # total
    tot = sum(float(s.get("signed_best_estimate") or 0) for s in signoff)
    ws.cell(r, 1, "TOTAL").font = BOLD
    tc = ws.cell(r, 2, tot); tc.font = BOLD; tc.number_format = MONEY
    _autosize(ws, [26, 20, 22, 26, 16, 18, 26])

    # sheet 2 — the selections behind them (factor by factor + rationale)
    ws2 = wb.create_sheet("Selections")
    _title(ws2, "Selected development patterns", "The governed factor pick each reserve rests on — with rationale and approver.", 6)
    _header_row(ws2, 4, ["Selection ID", "Line of business", "Source", "Tail", "Approved by", "Rationale"])
    r = 5
    for s in selections:
        ws2.cell(r, 1, s.get("selection_id"))
        ws2.cell(r, 2, (s.get("line_of_business_code") or "").replace("_", " "))
        ws2.cell(r, 3, (s.get("source_code") or "").replace("_", " ").title())
        ws2.cell(r, 4, float(s.get("tail_factor") or 0))
        ws2.cell(r, 5, s.get("approved_by") or "—")
        ws2.cell(r, 6, (s.get("rationale") or "")[:300])
        for c in range(1, 7):
            ws2.cell(r, c).border = BORDER
            ws2.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    _autosize(ws2, [26, 22, 20, 8, 22, 60])

    # sheet 3 — the audit trail (who/when/what)
    ws3 = wb.create_sheet("Audit trail")
    _title(ws3, "Audit trail", "Append-only record of every action behind this pack.", 5)
    _header_row(ws3, 4, ["When (UTC)", "Event", "Object", "Detail", "By"])
    r = 5
    for e in audit:
        ws3.cell(r, 1, (e.get("created_at") or "")[:19].replace("T", " "))
        ws3.cell(r, 2, (e.get("event_type") or "").replace("_", " "))
        ws3.cell(r, 3, e.get("entity_id") or "—")
        ws3.cell(r, 4, (e.get("detail") or "")[:200])
        ws3.cell(r, 5, e.get("actor") or "—")
        for c in range(1, 6):
            ws3.cell(r, c).border = BORDER
            ws3.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    _autosize(ws3, [20, 22, 24, 70, 26])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def writeback_template(lobs, n_factors=8):
    """Build the D2 write-back template: an analyst fills line, factors, tail, rationale;
    the app reads it back to land a PENDING_APPROVAL selection. Returns bytes."""
    wb = Workbook()
    ws = wb.active; ws.title = "Selection"
    _title(ws, "Reserving Workbench — selection write-back",
           "Fill the yellow cells and upload on the Engines page. It lands a PENDING_APPROVAL "
           "selection in the SAME governed table the app writes — it does not book anything.", 4)
    # instructions
    ws.cell(4, 1, "Line of business (code):").font = BOLD
    dv_note = ws.cell(4, 3, "one of: " + ", ".join(lobs)); dv_note.font = SUB
    ws.cell(5, 1, "Tail factor:").font = BOLD
    ws.cell(6, 1, "Rationale (required):").font = BOLD
    yellow = PatternFill("solid", fgColor="FEF3C7")
    for rr in (4, 5, 6):
        cell = ws.cell(rr, 2); cell.fill = yellow; cell.border = BORDER
    ws.cell(4, 2, lobs[0] if lobs else "COMMERCIAL_PROPERTY")
    ws.cell(5, 2, 1.01)
    ws.cell(6, 2, "")
    # factor grid
    _header_row(ws, 8, ["Development step"] + [f"{k}→{k+1}" for k in range(n_factors)])
    ws.cell(9, 1, "Selected factor").font = BOLD
    for k in range(n_factors):
        cell = ws.cell(9, 2 + k); cell.fill = yellow; cell.border = BORDER; cell.number_format = "0.000"
        cell.value = 1.0
    ws.cell(11, 1, "Note: an override away from the empirical pattern requires a rationale (≥10 chars), "
                   "exactly as in the app — the same governance, via a different door.").font = SUB
    ws.merge_cells(start_row=11, start_column=1, end_row=11, end_column=9)
    _autosize(ws, [22] + [10] * n_factors)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def parse_writeback(data: bytes):
    """Read a filled write-back template back into a selection dict. Returns
    {lob, tail, rationale, factors} or raises ValueError with a friendly message."""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise ValueError(f"couldn't read the workbook: {e}")
    ws = wb["Selection"] if "Selection" in wb.sheetnames else wb.active
    lob = ws.cell(4, 2).value
    tail = ws.cell(5, 2).value
    rationale = ws.cell(6, 2).value or ""
    if not lob:
        raise ValueError("line of business (cell B4) is empty")
    factors = []
    for k in range(8):
        v = ws.cell(9, 2 + k).value
        if v is None:
            break
        try:
            factors.append(round(float(v), 4))
        except (TypeError, ValueError):
            raise ValueError(f"factor {k}→{k+1} (row 9) is not a number: {v!r}")
    if len(factors) < 2:
        raise ValueError("need at least two development factors in row 9")
    try:
        tail = float(tail) if tail is not None else 1.01
    except (TypeError, ValueError):
        tail = 1.01
    return {"lob": str(lob).strip().upper(), "tail": tail,
            "rationale": str(rationale).strip(), "factors": factors}
